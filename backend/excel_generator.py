import io
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Border, Side, Alignment, numbers
)
from openpyxl.utils import get_column_letter

# ── Color palette (from original file) ──────────────────────────────────────
DARK_BLUE  = "1F3864"   # header investment
DARK_GREEN = "385623"   # header depreciation / costs production
BLUE_INPUT = "0070C0"   # user input cells (blue text)
WHITE      = "FFFFFF"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=None, size=11, name="Calibri"):
    kw = dict(bold=bold, size=size, name=name)
    if color:
        kw["color"] = color
    return Font(**kw)

def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

HEADER_BLUE  = {"fill": _fill(DARK_BLUE),  "font": _font(bold=True, color=WHITE), "border": _border(), "alignment": _align("center")}
HEADER_GREEN = {"fill": _fill(DARK_GREEN), "font": _font(bold=True, color=WHITE), "border": _border(), "alignment": _align("center")}
TOTAL_BLUE   = {"fill": _fill(DARK_BLUE),  "font": _font(bold=True, color=WHITE), "border": _border(), "alignment": _align("left")}
TOTAL_GREEN  = {"fill": _fill(DARK_GREEN), "font": _font(bold=True, color=WHITE), "border": _border(), "alignment": _align("left")}
INPUT_STYLE  = {"font": _font(bold=True, color=BLUE_INPUT), "border": _border(), "alignment": _align("right")}
NORMAL_STYLE = {"font": _font(), "border": _border(), "alignment": _align("left")}
RESULT_STYLE = {"font": _font(bold=True), "border": _border(), "alignment": _align("right")}

def _apply(cell, style: dict):
    for attr, val in style.items():
        setattr(cell, attr, val)

def _set(ws, row, col, value, style=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if style:
        _apply(cell, style)
    if number_format:
        cell.number_format = number_format
    return cell

CLP = '#,##0'
PCT = '0.00%'

# ─────────────────────────────────────────────────────────────────────────────
def build_sheet1(wb, inversiones):
    ws = wb.create_sheet("1. Inversión")
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 20

    # Row 2: title
    ws.merge_cells('B2:F2')
    c = ws['B2']
    c.value = "1. INVERSIÓN Y DEPRECIACIÓN "
    _apply(c, HEADER_BLUE)
    c.font = _font(bold=True, color=WHITE, size=13)

    # Row 3: instruction
    ws.merge_cells('B3:F3')
    c = ws['B3']
    c.value = 'En el ítem "Inversión" agregar las máquinas, Infraestructura y Equipos para la producción a escala comercial'
    _apply(c, {"font": _font(size=9), "alignment": _align(wrap=True)})

    ws.row_dimensions[3].height = 25

    # Row 5: section headers
    ws.merge_cells('B5:C5')
    _set(ws, 5, 2, 'INVERSIÓN', HEADER_BLUE)
    ws.merge_cells('D5:F5')
    _set(ws, 5, 4, 'Datos de Depreciación', HEADER_GREEN)

    # Row 6: column headers
    _set(ws, 6, 2, 'Maquinaria, Equipo o Infraestructura', HEADER_BLUE)
    _set(ws, 6, 3, 'MONTO ($)',          HEADER_BLUE)
    _set(ws, 6, 4, 'VIDA ÚTIL (años)',   HEADER_GREEN)
    _set(ws, 6, 5, 'Tasa de Depreciación', HEADER_GREEN)
    _set(ws, 6, 6, 'Costo de Depreciación', HEADER_GREEN)

    max_rows = max(len(inversiones), 11)
    start_row = 7

    for i in range(max_rows):
        r = start_row + i
        if i < len(inversiones):
            inv = inversiones[i]
            _set(ws, r, 2, inv.nombre, NORMAL_STYLE)
            _set(ws, r, 3, inv.monto, INPUT_STYLE, CLP)
            _set(ws, r, 4, inv.vida_util, INPUT_STYLE)
        else:
            _set(ws, r, 2, None, NORMAL_STYLE)
            _set(ws, r, 3, None, {"border": _border()})
            _set(ws, r, 4, None, {"border": _border()})
        _set(ws, r, 5, f'=IFERROR(1/D{r},0)', {"border": _border(), "alignment": _align("right")}, PCT)
        _set(ws, r, 6, f'=IFERROR(C{r}*E{r},0)', {"border": _border(), "alignment": _align("right")}, CLP)

    total_row = start_row + max_rows
    ws.merge_cells(f'D{total_row}:E{total_row}')
    _set(ws, total_row, 2, 'TOTAL INVERSION', TOTAL_BLUE)
    _set(ws, total_row, 3, f'=SUM(C{start_row}:C{total_row-1})', TOTAL_BLUE, CLP)
    _set(ws, total_row, 4, 'COSTO DEPRECIACIÓN', TOTAL_GREEN)
    _set(ws, total_row, 6, f'=SUM(F{start_row}:F{total_row-1})', TOTAL_GREEN, CLP)

    ws.sheet_view.showGridLines = False
    return total_row  # returns row of totals so other sheets can reference

def build_sheet2(wb, costos_fijos, costos_variables):
    ws = wb.create_sheet("2. Costos")
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 30

    # Title
    ws.merge_cells('B2:C2')
    c = ws['B2']
    c.value = "2. COSTOS "
    _apply(c, HEADER_BLUE)
    c.font = _font(bold=True, color=WHITE, size=13)

    # Costos Fijos section
    ws.merge_cells('B5:C5')
    _set(ws, 5, 2, 'COSTOS FIJOS', HEADER_BLUE)
    _set(ws, 5, 4, '', {"fill": _fill(DARK_BLUE)})

    _set(ws, 6, 2, 'Tipo de costos Fijos',   HEADER_BLUE)
    _set(ws, 6, 3, 'Costo Mensual ($)',       HEADER_BLUE)
    _set(ws, 6, 4, 'Observación',             HEADER_BLUE)

    max_fijos = max(len(costos_fijos), 9)
    start_fijos = 7
    for i in range(max_fijos):
        r = start_fijos + i
        if i < len(costos_fijos):
            cf = costos_fijos[i]
            _set(ws, r, 2, cf.tipo, NORMAL_STYLE)
            _set(ws, r, 3, cf.costo_mensual, INPUT_STYLE, CLP)
            _set(ws, r, 4, cf.observacion or '', {"border": _border(), "alignment": _align(wrap=True)})
        else:
            for c in [2, 3, 4]:
                _set(ws, r, c, None, {"border": _border()})

    total_fijos_row = start_fijos + max_fijos
    total_anual_row = total_fijos_row + 1

    _set(ws, total_fijos_row, 2, 'TOTAL COSTOS FIJOS MENSUAL',
         TOTAL_BLUE)
    _set(ws, total_fijos_row, 3,
         f'=SUM(C{start_fijos}:C{total_fijos_row-1})', TOTAL_BLUE, CLP)
    _set(ws, total_fijos_row, 4, '', {"fill": _fill(DARK_BLUE)})

    _set(ws, total_anual_row, 2, 'TOTAL COSTOS FIJOS ANUAL', TOTAL_BLUE)
    _set(ws, total_anual_row, 3, f'=C{total_fijos_row}*12', TOTAL_BLUE, CLP)
    _set(ws, total_anual_row, 4, '', {"fill": _fill(DARK_BLUE)})

    # Costos Variables section
    cv_start = total_anual_row + 2
    ws.merge_cells(f'B{cv_start}:C{cv_start}')
    _set(ws, cv_start, 2, 'COSTOS DE PRODUCCIÓN', HEADER_GREEN)
    _set(ws, cv_start, 4, '', {"fill": _fill(DARK_GREEN)})

    _set(ws, cv_start+1, 2, 'Tipo de costos Variable', HEADER_GREEN)
    _set(ws, cv_start+1, 3, 'Costo ($)',               HEADER_GREEN)
    _set(ws, cv_start+1, 4, 'Observación',             HEADER_GREEN)

    max_var = max(len(costos_variables), 10)
    start_var = cv_start + 2
    for i in range(max_var):
        r = start_var + i
        if i < len(costos_variables):
            cv = costos_variables[i]
            _set(ws, r, 2, cv.tipo, NORMAL_STYLE)
            _set(ws, r, 3, cv.costo, INPUT_STYLE, CLP)
            _set(ws, r, 4, cv.observacion or '', {"border": _border(), "alignment": _align(wrap=True)})
        else:
            for c in [2, 3, 4]:
                _set(ws, r, c, None, {"border": _border()})

    total_var_row = start_var + max_var
    _set(ws, total_var_row, 2, 'COSTO PRODUCCIÓN UNIDAD',
         TOTAL_GREEN)
    _set(ws, total_var_row, 3,
         f'=SUM(C{start_var}:C{total_var_row-1})', TOTAL_GREEN, CLP)
    _set(ws, total_var_row, 4, '', {"fill": _fill(DARK_GREEN)})

    ws.sheet_view.showGridLines = False
    return total_fijos_row, total_anual_row, total_var_row

def build_sheet3(wb, data, inv_total_row, fijos_anual_row, var_total_row):
    ws = wb.create_sheet("3. Flujo de Caja")
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    for col in ['D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['I'].width = 3
    ws.column_dimensions['J'].width = 40

    # Title
    ws.merge_cells('B2:H2')
    c = ws['B2']
    c.value = "3. FLUJO DE CAJA"
    _apply(c, HEADER_BLUE)
    c.font = _font(bold=True, color=WHITE, size=13)

    # Datos clave section
    _set(ws, 4, 2, 'Datos Clave',    HEADER_BLUE)
    _set(ws, 4, 3, 'Monto ($)',      HEADER_BLUE)
    ws.merge_cells('D4:H4')
    _set(ws, 4, 4, 'Observación',   HEADER_BLUE)

    _set(ws, 5, 2, 'Valor de Venta',   NORMAL_STYLE)
    _set(ws, 5, 3, data.precio_venta,  INPUT_STYLE, CLP)
    ws.merge_cells('D5:H5')
    _set(ws, 5, 4, 'Valor venta de unidad de producto', {"font": _font(size=9), "alignment": _align(wrap=True)})

    _set(ws, 6, 2, 'Costo producción', NORMAL_STYLE)
    _set(ws, 6, 3, f"='2. Costos'!C{var_total_row}", INPUT_STYLE, CLP)
    ws.merge_cells('D6:H6')
    _set(ws, 6, 4, 'Costo producción de una unidad de producto', {"font": _font(size=9), "alignment": _align(wrap=True)})

    # Flujo de Caja table header
    ws.merge_cells('B8:H8')
    _set(ws, 8, 2, ' FLUJO DE CAJA', HEADER_BLUE)

    _set(ws, 9, 2, 'ITEM',  HEADER_BLUE)
    ws.merge_cells('C9:H9')
    _set(ws, 9, 3, 'Año',   HEADER_BLUE)

    years = [0, 1, 2, 3, 4, 5]
    year_cols = {'C': 0, 'D': 1, 'E': 2, 'F': 3, 'G': 4, 'H': 5}
    _set(ws, 10, 2, None, {"fill": _fill(DARK_BLUE)})
    for col_letter, yr in zip(['C','D','E','F','G','H'], years):
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 10, col_idx, yr, HEADER_BLUE)

    ws.merge_cells('J10:N10')
    _set(ws, 10, 10, 'Observación', HEADER_BLUE)

    # Row 11: INVERSIÓN
    _set(ws, 11, 2, 'INVERSIÓN', NORMAL_STYLE)
    _set(ws, 11, 3, f"=IFERROR('1. Inversión'!C{inv_total_row},0)", RESULT_STYLE, CLP)
    for col in [4, 5, 6, 7, 8]:
        _set(ws, 11, col, '-', {"border": _border(), "alignment": _align("center")})
    ws.merge_cells('J11:N11')
    _set(ws, 11, 10, 'La inversión se considera solamente en el año 0', {"font": _font(size=9), "alignment": _align(wrap=True)})

    # Row 12: Cantidad de venta
    _set(ws, 12, 2, 'Cantidad de venta', NORMAL_STYLE)
    _set(ws, 12, 3, '-', {"border": _border(), "alignment": _align("center")})
    cantidades = (data.cantidades_venta + [0]*5)[:5]
    for i, (col_letter) in enumerate(['D','E','F','G','H']):
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 12, col_idx, cantidades[i], INPUT_STYLE, '#,##0')
    ws.merge_cells('J12:N12')
    _set(ws, 12, 10, 'Cantidad de unidades vendidas al año según estimación de demanda.', {"font": _font(size=9), "alignment": _align(wrap=True)})

    # Row 13: INGRESO BRUTO
    _set(ws, 13, 2, 'INGRESO BRUTO', NORMAL_STYLE)
    _set(ws, 13, 3, '-', {"border": _border(), "alignment": _align("center")})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 13, col_idx, f'=IFERROR({col_letter}12*$C$5,0)', RESULT_STYLE, CLP)
    ws.merge_cells('J13:N13')
    _set(ws, 13, 10, 'Ingresos = Precio de Venta x Unidades anuales vendidas', {"font": _font(size=9), "alignment": _align(wrap=True)})

    # Row 14: Costo Variable
    _set(ws, 14, 2, 'Costo Variable', NORMAL_STYLE)
    _set(ws, 14, 3, '-', {"border": _border(), "alignment": _align("center")})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 14, col_idx, f'=IFERROR({col_letter}12*$C$6,0)', RESULT_STYLE, CLP)
    ws.merge_cells('J14:N14')
    _set(ws, 14, 10, 'Costo Variable = Costo de producción de unidad x precio de venta', {"font": _font(size=9), "alignment": _align(wrap=True)})

    # Row 15: Costo Fijo
    _set(ws, 15, 2, 'Costo Fijo', NORMAL_STYLE)
    _set(ws, 15, 3, '-', {"border": _border(), "alignment": _align("center")})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 15, col_idx, f"=IFERROR('2. Costos'!$C${fijos_anual_row},0)", RESULT_STYLE, CLP)
    ws.merge_cells('J15:N15')
    _set(ws, 15, 10, 'Costos Fijos Anuales', {"font": _font(size=9)})

    # Row 16: Depreciación
    _set(ws, 16, 2, 'Depreciación', NORMAL_STYLE)
    _set(ws, 16, 3, '-', {"border": _border(), "alignment": _align("center")})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 16, col_idx, f"=IFERROR('1. Inversión'!$F${inv_total_row},0)", RESULT_STYLE, CLP)
    ws.merge_cells('J16:N16')
    _set(ws, 16, 10, 'Costos de Depreciación', {"font": _font(size=9)})

    # Row 17: Total Egresos
    _set(ws, 17, 2, 'Total de Egresos', NORMAL_STYLE)
    _set(ws, 17, 3, None, {"border": _border()})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 17, col_idx, f'=IFERROR(SUM({col_letter}14:{col_letter}16),0)', RESULT_STYLE, CLP)
    ws.merge_cells('J17:N17')
    _set(ws, 17, 10, 'Total de Egresos (suma de costos)', {"font": _font(size=9)})

    # Row 18: UTILIDAD
    _set(ws, 18, 2, 'UTILIDAD', NORMAL_STYLE)
    _set(ws, 18, 3, '-', {"border": _border(), "alignment": _align("center")})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 18, col_idx, f'=IFERROR({col_letter}13-{col_letter}17,0)', RESULT_STYLE, CLP)
    ws.merge_cells('J18:N18')
    _set(ws, 18, 10, 'Utilidad Antes de los Impuestos', {"font": _font(size=9)})

    # Row 19: Impuesto
    _set(ws, 19, 2, 'Impuesto (27%)', NORMAL_STYLE)
    _set(ws, 19, 3, '-', {"border": _border(), "alignment": _align("center")})
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 19, col_idx, f'=IFERROR(IF({col_letter}18>0,{col_letter}18*0.27,0),0)', RESULT_STYLE, CLP)
    ws.merge_cells('J19:N19')
    _set(ws, 19, 10, 'Impuesto de Primera Categoría (27%)', {"font": _font(size=9)})

    # Row 20: FLUJO DE CAJA
    fc_style = {**TOTAL_BLUE}
    _set(ws, 20, 2, 'FLUJO DE CAJA', TOTAL_BLUE)
    _set(ws, 20, 3, '=-C11', {**TOTAL_BLUE, "number_format": CLP}, CLP)
    for col_letter in ['D','E','F','G','H']:
        col_idx = ord(col_letter) - ord('A') + 1
        _set(ws, 20, col_idx, f'=IFERROR({col_letter}18-{col_letter}19+{col_letter}16,0)', TOTAL_BLUE, CLP)
    ws.merge_cells('J20:N20')
    _set(ws, 20, 10, 'Flujo de Caja Final (Inversión siempre se considera en negativo)', {"font": _font(size=9)})

    # Indicators section
    ws.merge_cells('B22:H22')
    _set(ws, 22, 2, 'INDICADORES DE EVALUACIÓN', HEADER_BLUE)
    _set(ws, 22, 3, None, {"fill": _fill(DARK_BLUE)})

    _set(ws, 23, 2, 'INDICADOR',    HEADER_BLUE)
    _set(ws, 23, 3, 'RESULTADO',    HEADER_BLUE)
    ws.merge_cells('D23:H23')
    _set(ws, 23, 4, 'EVALUACIÓN',   HEADER_BLUE)

    _set(ws, 24, 2, 'Tasa de Interés', NORMAL_STYLE)
    _set(ws, 24, 3, data.tasa_interes, INPUT_STYLE, PCT)
    ws.merge_cells('D24:H24')
    _set(ws, 24, 4, 'Normalmente se considera tasa del 12%', {"font": _font(size=9)})

    _set(ws, 25, 2, 'VAN', NORMAL_STYLE)
    _set(ws, 25, 3, '=IFERROR(NPV(C24,D20:H20)+C20,0)', RESULT_STYLE, CLP)
    ws.merge_cells('D25:H25')
    _set(ws, 25, 4, 'VAN > 0; proyecto rentable. VAN < 0; proyecto no rentable.', {"font": _font(size=9)})

    _set(ws, 26, 2, 'TIR', NORMAL_STYLE)
    _set(ws, 26, 3, '=IFERROR(IRR(C20:H20,C24),"N/A")', RESULT_STYLE, PCT)
    ws.merge_cells('D26:H26')
    _set(ws, 26, 4, 'Si TIR > Tasa interés; proyecto rentable.', {"font": _font(size=9)})

    _set(ws, 27, 2, 'Payback (año)', NORMAL_STYLE)
    _set(ws, 27, 3, "=IFERROR(IF(ISNUMBER('4. Payback'!B14),'4. Payback'!B14,\"No se recupera\"),0)", RESULT_STYLE, '0.00')
    ws.merge_cells('D27:H27')
    _set(ws, 27, 4, 'Indica los años en donde se recupera lo invertido.', {"font": _font(size=9)})

    _set(ws, 28, 2, 'Costo/Beneficio', NORMAL_STYLE)
    _set(ws, 28, 3, "=IFERROR('5. C-B'!C20,0)", RESULT_STYLE, '0.00')
    ws.merge_cells('D28:H28')
    _set(ws, 28, 4, 'Si C/B > 1; proyecto rentable. Si C/B < 1; proyecto no rentable.', {"font": _font(size=9)})

    ws.sheet_view.showGridLines = False

def build_sheet4(wb):
    ws = wb.create_sheet("4. Payback")
    ws.column_dimensions['A'].width = 30
    for col in ['B','C','D','E','F','G']:
        ws.column_dimensions[col].width = 14
    ws.column_dimensions['H'].width = 3
    ws.column_dimensions['I'].width = 40

    _set(ws, 1, 1, 'CÁLCULO AUTOMÁTICO DE PAYBACK', HEADER_BLUE)
    ws.merge_cells('A1:G1')
    _set(ws, 2, 1, 'El payback indica en cuánto tiempo se recupera la inversión inicial. Se calcula usando el flujo de caja acumulado.',
         {"font": _font(size=9), "alignment": _align(wrap=True)})
    ws.merge_cells('A2:G2')
    ws.row_dimensions[2].height = 25

    for i, label in enumerate(['Año','0','1','2','3','4','5']):
        _set(ws, 3, i+1, label, HEADER_BLUE)

    _set(ws, 4, 1, 'Flujo de Caja', NORMAL_STYLE)
    for i, col in enumerate(['C','D','E','F','G','H']):
        col_idx = i + 2
        _set(ws, 4, col_idx, f"='3. Flujo de Caja'!{col}20", RESULT_STYLE, CLP)

    _set(ws, 5, 1, 'Flujo Acumulado', NORMAL_STYLE)
    _set(ws, 5, 2, '=B4', RESULT_STYLE, CLP)
    for i in range(1, 6):
        col_prev = get_column_letter(i + 1)
        col_curr = get_column_letter(i + 2)
        _set(ws, 5, i+2, f'={col_curr}4+{col_prev}5', RESULT_STYLE, CLP)

    _set(ws, 7, 1, 'CÁLCULO PASO A PASO', HEADER_BLUE)
    ws.merge_cells('A7:G7')

    calcs = [
        ('Periodos con acumulado negativo', '=COUNTIF(B5:G5,"<0")', 'Cuenta cuántos años (incl. año 0) con flujo acumulado negativo.'),
        ('Años completos para recuperar',   '=MAX(0,B8-1)',          'Años enteros antes de que el acumulado deje de ser negativo.'),
        ('Monto faltante al final del año', '=IFERROR(ABS(INDEX(B5:G5,B8)),0)', 'Cuánto falta por recuperar al cierre del último año negativo.'),
        ('Flujo del año siguiente',         '=IFERROR(INDEX(B4:G4,B8+1),0)',    'Flujo de caja del año en que se recupera la inversión.'),
        ('Fracción de año adicional',        '=IFERROR(B10/B11,0)',              'Porción del año siguiente para completar la recuperación.'),
    ]
    for i, (label, formula, obs) in enumerate(calcs):
        r = 8 + i
        _set(ws, r, 1, label, NORMAL_STYLE)
        _set(ws, r, 2, formula, RESULT_STYLE, '0.00' if i == 4 else CLP)
        _set(ws, r, 9, obs, {"font": _font(size=9), "alignment": _align(wrap=True)})

    _set(ws, 13, 1, 'PAYBACK (años)', TOTAL_BLUE)
    _set(ws, 13, 2, '=IF(G5<0,"No se recupera",IFERROR(B9+B12,"N/A"))', TOTAL_BLUE, '0.00')

    _set(ws, 14, 1, 'Años enteros', NORMAL_STYLE)
    _set(ws, 14, 2, '=IFERROR(INT(B13),"")', RESULT_STYLE)
    _set(ws, 15, 1, 'Meses', NORMAL_STYLE)
    _set(ws, 15, 2, '=IFERROR(INT((B13-B14)*12),"")', RESULT_STYLE)
    _set(ws, 16, 1, 'Días', NORMAL_STYLE)
    _set(ws, 16, 2, '=IFERROR(ROUND(((B13-B14)*12-B15)*30,0),"")', RESULT_STYLE)

    _set(ws, 17, 1, 'Resultado en lenguaje natural:', NORMAL_STYLE)
    _set(ws, 17, 2, '=IF(ISNUMBER(B13),B14&" año(s), "&B15&" mes(es) y "&B16&" día(s)","No se recupera la inversión en 5 años")',
         {"font": _font(bold=True), "border": _border()})
    ws.merge_cells('B17:G17')

    _set(ws, 18, 1, 'Evaluación', HEADER_GREEN)
    _set(ws, 18, 2, '=IF(ISNUMBER(B13),IF(B13<5,"El proyecto recupera la inversión dentro del horizonte de 5 años.","El payback excede el horizonte de evaluación."),"⚠ Los flujos no logran recuperar la inversión en 5 años")',
         HEADER_GREEN)
    ws.merge_cells('B18:G18')

    ws.sheet_view.showGridLines = False

def build_sheet5(wb, fijos_anual_row, var_total_row, tasa_cb):
    ws = wb.create_sheet("5. C-B")
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 16
    for col in ['D','E','F','G','H']:
        ws.column_dimensions[col].width = 14

    ws.merge_cells('B2:H2')
    c = ws['B2']
    c.value = "5. COSTO/BENEFICIO"
    _apply(c, HEADER_BLUE)
    c.font = _font(bold=True, color=WHITE, size=13)

    ws.merge_cells('B4:H4')
    _set(ws, 4, 2, ' FLUJO DE CAJA', HEADER_BLUE)

    _set(ws, 5, 2, 'ITEM', HEADER_BLUE)
    ws.merge_cells('C5:H5')
    _set(ws, 5, 3, 'Año', HEADER_BLUE)

    for i, yr in enumerate([0,1,2,3,4,5]):
        _set(ws, 6, i+3 if i > 0 else 3, yr, HEADER_BLUE)
    _set(ws, 6, 2, None, {"fill": _fill(DARK_BLUE)})
    for i, yr in enumerate([0,1,2,3,4,5]):
        col_idx = 3 + i
        _set(ws, 6, col_idx, yr, HEADER_BLUE)

    rows = [
        ('INVERSIÓN',    "='3. Flujo de Caja'!C11", ['-']*5),
        ('INGRESO BRUTO','-', [f"='3. Flujo de Caja'!{c}13" for c in ['D','E','F','G','H']]),
        ('Costo Variable','-',[f"='3. Flujo de Caja'!{c}14" for c in ['D','E','F','G','H']]),
        ('Costo Fijo',   '-', [f"='3. Flujo de Caja'!{c}15" for c in ['D','E','F','G','H']]),
        ('Depreciación', '-', [f"='3. Flujo de Caja'!{c}16" for c in ['D','E','F','G','H']]),
        ('Impuesto (27%)','-',[f"='3. Flujo de Caja'!{c}19" for c in ['D','E','F','G','H']]),
    ]
    for i, (label, c0, rest) in enumerate(rows):
        r = 7 + i
        _set(ws, r, 2, label, NORMAL_STYLE)
        _set(ws, r, 3, c0, {"border": _border(), "alignment": _align("center" if c0 == '-' else "right")},
             CLP if c0 != '-' else None)
        for j, val in enumerate(rest):
            _set(ws, r, 4+j, val, {"border": _border(), "alignment": _align("center" if val == '-' else "right")},
                 CLP if val != '-' else None)

    total_row = 13
    _set(ws, total_row, 2, 'TOTAL EGRESOS', TOTAL_GREEN)
    _set(ws, total_row, 3, '-', {"border": _border(), "fill": _fill(DARK_GREEN)})
    for i, col in enumerate(['D','E','F','G','H']):
        col_idx = 4 + i
        _set(ws, total_row, col_idx, f'=IFERROR(SUM({col}9:{col}12),0)', TOTAL_GREEN, CLP)

    ws.merge_cells('B15:H15')
    _set(ws, 15, 2, 'EVALUACIÓN', HEADER_BLUE)

    _set(ws, 16, 2, 'INDICADOR', HEADER_BLUE)
    _set(ws, 16, 3, 'RESULTADO', HEADER_BLUE)
    ws.merge_cells('D16:H16')

    _set(ws, 17, 2, 'Tasa de Interés', NORMAL_STYLE)
    _set(ws, 17, 3, tasa_cb, INPUT_STYLE, PCT)

    _set(ws, 18, 2, 'VAN INGRESOS', NORMAL_STYLE)
    _set(ws, 18, 3, '=IFERROR(NPV(C17,D8:H8),0)', RESULT_STYLE, CLP)

    _set(ws, 19, 2, 'VAN EGRESOS', NORMAL_STYLE)
    _set(ws, 19, 3, '=IFERROR(NPV(C17,D13:H13),0)', RESULT_STYLE, CLP)

    _set(ws, 20, 2, 'Egresos + Inversión', NORMAL_STYLE)
    _set(ws, 20, 3, '=IFERROR(C19+C7,0)', RESULT_STYLE, CLP)

    _set(ws, 21, 2, 'Costo/Beneficio', TOTAL_BLUE)
    _set(ws, 21, 3, '=IFERROR(C18/C20,0)', TOTAL_BLUE, '0.00')
    ws.merge_cells('D21:H21')
    _set(ws, 21, 4, 'Si C/B > 1; proyecto rentable. Si C/B < 1; proyecto no rentable.', {"font": _font(size=9)})

    ws.sheet_view.showGridLines = False

def build_sheet6(wb, modelo):
    ws = wb.create_sheet("6. Diagrama Modelo Negocios")

    col_widths = {'A': 3, 'B': 3, 'C': 22, 'D': 22, 'E': 22, 'F': 3, 'G': 22, 'H': 22}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    canvas_fill_blue  = _fill("BDD7EE")
    canvas_fill_green = _fill("E2EFDA")
    canvas_fill_gray  = _fill("F2F2F2")

    def canvas_cell(ws, row, col, value, fill, merge_end_col=None):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = fill
        c.font = Font(name="Calibri", size=10)
        c.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        return c

    def canvas_header(ws, row, col, value, merge_end_col=None):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = _fill(DARK_BLUE)
        c.font = Font(name="Calibri", bold=True, color=WHITE, size=10)
        c.border = Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        )
        c.alignment = Alignment(horizontal='center', vertical='center')
        return c

    m = modelo or object()
    def g(attr): return getattr(m, attr, '') or ''

    # Row 1: Title
    ws.merge_cells('B1:H1')
    t = ws['B1']
    t.value = "6. DIAGRAMA MODELO DE NEGOCIOS (CANVAS)"
    t.fill = _fill(DARK_BLUE)
    t.font = Font(name="Calibri", bold=True, color=WHITE, size=13)
    t.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    # Headers row 3
    headers = [('C', 'Socios Estratégicos'), ('D', 'Recursos Clave'), ('E', 'Propuesta de Valor'),
               ('G', 'Relación con el Cliente'), ('H', 'Segmento de Clientes')]
    for col_l, hdr in headers:
        col = ord(col_l) - ord('A') + 1
        canvas_header(ws, 3, col, hdr)

    canvas_header(ws, 4, 4, 'Actividades Clave')
    canvas_header(ws, 4, 7, 'Canales')

    # Content rows
    for row_heights in [(3, 60), (4, 60), (5, 60), (6, 60), (7, 50), (8, 50)]:
        ws.row_dimensions[row_heights[0]].height = row_heights[1]

    fields = {
        (3, 3): g('socios_estrategicos'), (3, 4): g('recursos_clave'),
        (3, 5): g('propuesta_valor'),     (3, 7): g('relacion_cliente'),
        (3, 8): g('segmento_clientes'),
        (4, 4): g('actividades_clave'),   (4, 5): g('propuesta_valor'),
        (4, 7): g('canales'),             (4, 8): g('segmento_clientes'),
    }

    # Estructura de costos y flujo de ingresos
    canvas_header(ws, 6, 3, 'Estructura de Costos')
    ws.merge_cells('C6:E6')
    canvas_header(ws, 6, 7, 'Flujo de Ingresos')
    ws.merge_cells('G6:H6')

    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 60

    c1 = ws.cell(row=7, column=3, value=g('estructura_costos'))
    c1.fill = canvas_fill_blue; c1.font = Font(size=10); c1.alignment = Alignment(wrap_text=True, vertical='top')
    c1.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    ws.merge_cells('C7:E7')

    c2 = ws.cell(row=7, column=7, value=g('flujo_ingresos'))
    c2.fill = canvas_fill_green; c2.font = Font(size=10); c2.alignment = Alignment(wrap_text=True, vertical='top')
    c2.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    ws.merge_cells('G7:H7')

    # Barreras de entrada
    canvas_header(ws, 9, 3, 'Barreras de Entrada')
    ws.merge_cells('C9:H9')
    ws.row_dimensions[9].height = 20

    c3 = ws.cell(row=10, column=3, value=g('barreras_entrada'))
    c3.fill = canvas_fill_gray; c3.font = Font(size=10); c3.alignment = Alignment(wrap_text=True, vertical='top')
    c3.border = Border(left=Side(style='medium'), right=Side(style='medium'), top=Side(style='medium'), bottom=Side(style='medium'))
    ws.merge_cells('C10:H10')
    ws.row_dimensions[10].height = 60

    ws.sheet_view.showGridLines = False

# ─────────────────────────────────────────────────────────────────────────────
def generate_excel(data) -> io.BytesIO:
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    inv_total_row = build_sheet1(wb, data.inversiones)
    fijos_row, anual_row, var_row = build_sheet2(wb, data.costos_fijos, data.costos_variables)
    build_sheet3(wb, data, inv_total_row, anual_row, var_row)
    build_sheet4(wb)
    build_sheet5(wb, anual_row, var_row, data.tasa_cb)
    build_sheet6(wb, data.modelo_negocio)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
